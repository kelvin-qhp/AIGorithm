import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime, date
import json

class ExcelExporter:
    """高级Excel导出工具"""
    
    def __init__(self):
        self.workbook = None
        self.current_sheet = None

    def export_list(self, data, filename, sheet_name='Sheet1',
                         include_index=False, auto_width=True):
        df = pd.DataFrame(data)
        self.export_dataframe(df=df,filename=filename,sheet_name=sheet_name,include_index=include_index,auto_width=auto_width)

    def export_dataframe(self, df, filename, sheet_name='Sheet1', 
                        include_index=False, auto_width=True):
        """导出DataFrame到Excel"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=include_index)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 设置样式
            self._style_excel(worksheet, df, include_index)
            
            # 自动调整列宽
            if auto_width:
                self._auto_adjust_width(worksheet)
                
        print(f"DataFrame已导出到: {filename}")
        return filename
        
    def export_multiple_sheets(self, data_dict, filename):
        """导出多个sheet到同一个Excel文件"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                self._auto_adjust_width(worksheet)
                
        print(f"多个Sheet已导出到: {filename}")
        return filename
        
    def export_with_formatting(self, data, headers, filename, 
                               sheet_name='Sheet1', date_columns=None):
        """导出带格式化的数据"""
        self.workbook = openpyxl.Workbook()
        self.current_sheet = self.workbook.active
        self.current_sheet.title = sheet_name
        
        # 设置表头
        self._set_formatted_headers(headers)
        
        # 写入数据
        self._write_formatted_data(data, headers, date_columns)
        
        # 保存文件
        self.workbook.save(filename)
        print(f"格式化数据已导出到: {filename}")
        return filename
        
    def _set_formatted_headers(self, headers):
        """设置带样式的表头"""
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = self.current_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # 设置首行行高
        self.current_sheet.row_dimensions[1].height = 30
        
    def _write_formatted_data(self, data, headers, date_columns=None):
        """写入带格式的数据"""
        date_columns = date_columns or []
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = self.current_sheet.cell(row=row_idx, column=col_idx)
                
                # 处理日期格式
                if col_idx - 1 in date_columns and isinstance(value, (datetime, date)):
                    cell.value = value
                    cell.number_format = 'yyyy-mm-dd'
                elif isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = value
                    
                # 设置对齐和边框
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
    def _auto_adjust_width(self, worksheet):
        """自动调整列宽"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    def _style_excel(self, worksheet, df, include_index):
        """给Excel添加样式"""
        # 设置表头样式
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # 设置数据样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
        # 自动调整列宽
        self._auto_adjust_width(worksheet)

# 使用示例
def advanced_export_example():
    """高级导出示例"""
    
    # 示例1：导出DataFrame
    df = pd.DataFrame({
        '姓名': ['张三', '李四', '王五'],
        '年龄': [28, 32, 25],
        '工资': [15000, 18000, 12000],
        '入职日期': pd.to_datetime(['2023-01-15', '2022-06-20', '2023-08-10'])
    })
    
    exporter = ExcelExporter()
    exporter.export_dataframe(df, 'dataframe_export.xlsx')
    
    # 示例2：导出多个Sheet
    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    df2 = pd.DataFrame({'X': [7, 8, 9], 'Y': [10, 11, 12]})
    data_dict = {'Sheet1': df1, 'Sheet2': df2}
    exporter.export_multiple_sheets(data_dict, 'multi_sheet_export.xlsx')
    
    # 示例3：导出带格式的数据
    data = [
        ['张三', 28, 15000, datetime(2023, 1, 15)],
        ['李四', 32, 18000, datetime(2022, 6, 20)],
        ['王五', 25, 12000, datetime(2023, 8, 10)]
    ]
    headers = ['姓名', '年龄', '工资', '入职日期']
    exporter.export_with_formatting(data, headers, 'formatted_export.xlsx', date_columns=[3])

if __name__ == '__main__':
    advanced_export_example()